#!/usr/bin/python3

import ipaddress
import glob
import re
import openpyxl


def parse_line(line):
    if "ip address" in line:
        match = re.search(r"\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b", line)
        if match:
            return ipaddress.IPv4Interface(match.groups())
    return None


def print_as_table(ip_list):
    col_width = 17
    print(f'{"Сеть":{col_width}}{"Маска":{col_width}}')
    for ip in ip_list:
        print(f'{str(ip.network.network_address):{col_width}}{str(ip.netmask):{col_width}}')
    print(f'{len(ip_list)} unique networks')


def save_as_excel(ip_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "Сеть"
    ws['A1'].font = openpyxl.styles.Font(bold=True)
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    ws['B1'] = "Маска"
    ws['B1'].font = openpyxl.styles.Font(bold=True)
    ws['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    for n, ip in enumerate(ip_list):
        ws[f'A{str(n + 2)}'] = str(ip.network.network_address)
        ws[f'B{str(n + 2)}'] = str(ip.netmask)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    wb.save('addrplan.xlsx')


ips = set()
for f_name in glob.glob("/home/pm/tmp/config_files/*.log"):
    with open(f_name) as f:
        for cur_line in f:
            result = parse_line(cur_line)
            if result:
                ips.add(result)
print_as_table(ips)
save_as_excel(ips)

